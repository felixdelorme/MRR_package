import os
import netCDF4
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image

def extract_dates_from_netcdf(netcdf_file):
    ds = netCDF4.Dataset(netcdf_file)
    time = ds.variables['time']
    start = pd.to_datetime(time[:], unit='s')[0]
    end = pd.to_datetime(time[:], unit='s')[-1]
    ds.close()
    return start, end

def resize_image(image_path, output_path, base_width):
    img = Image.open(image_path)
    w_percent = (base_width / float(img.size[0]))
    h_size = int((float(img.size[1]) * float(w_percent)))
    img = img.resize((base_width, h_size), Image.Resampling.LANCZOS)
    img.save(output_path)

def clean_resized_images(directories):
    for directory in directories:
        for filename in os.listdir(directory):
            if filename.endswith('_resized.png'):
                os.remove(os.path.join(directory, filename))

def create_excel_file(directory, additional_directories, excel_file):
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Debut'
    ws['B1'] = 'Fin'
    ws['C1'] = 'Image du MRR'
    ws['D1'] = 'Température Moyenne'
    ws['E1'] = 'Précipitations et Albedo'
    ws['F1'] = 'Précipitations et la Température Moyenne'
    ws['G1'] = 'Albedo Journalier'
    #Ajouter des cellules au besoin 

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 50

    row = 2

    netcdf_files = sorted([f for f in os.listdir(directory) if f.endswith('.nc')])
    additional_images = [sorted(os.listdir(d)) for d in additional_directories]

    for filename in netcdf_files:
        netcdf_path = os.path.join(directory, filename)
        start_date, end_date = extract_dates_from_netcdf(netcdf_path)
        
        ws.cell(row=row, column=1, value=start_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=2, value=end_date.strftime('%Y-%m-%d %H:%M:%S'))

        main_img_filename = filename.replace('.nc', '.png')
        main_img_path = os.path.join(directory, main_img_filename)

        # Acces aux données du MRR
        if os.path.exists(main_img_path):
            resized_img_path = main_img_path.replace('.png', '_resized.png')
            resize_image(main_img_path, resized_img_path, 200)

            img = OpenpyxlImage(resized_img_path)
            ws.add_image(img, f'C{row}')
            ws.row_dimensions[row].height = img.height * 0.75
        else:
            ws.cell(row=row, column=3, value="donnée non voulu/ donnée non exploité")

        # Acces aux données des .Dat 
        for j, img_col in enumerate(['D', 'E', 'F', 'G']):
            corresponding_image_found = False
            for additional_img in additional_images[j]:
                additional_img_date_str = additional_img.split('_')[0]
                additional_img_date = pd.to_datetime(additional_img_date_str, format='%Y-%m-%d')
                if start_date.date() == additional_img_date.date():
                    additional_img_path = os.path.join(additional_directories[j], additional_img)
                    if os.path.exists(additional_img_path):
                        resized_img_path = additional_img_path.replace('.png', '_resized.png')
                        resize_image(additional_img_path, resized_img_path, 375)

                        img = OpenpyxlImage(resized_img_path)
                        ws.add_image(img, f'{img_col}{row}')
                        corresponding_image_found = True
                        break
            
            if not corresponding_image_found:
                ws.cell(row=row, column=j+4, value="donnée non voulu/ donnée non exploité")

        row += 1

    wb.save(excel_file)

    clean_resized_images([directory] + additional_directories) # Supp les images reformater pour l'excel. 

# PATH DU MRR et des differents .DAT => à modifier selon vos PATH
directory = '/Users/felixdelorme/ESISAR/Stage/stage_felix_2024_MRR_Bolivie/MK_packaged_v2/MK_processed/202311'  # Update with the actual path
additional_directories = [
    '/Users/felixdelorme/ESISAR/Stage/stage_felix_2024_MRR_Bolivie/MK_packaged_v2/Dat processing/temperature_figures', 
    '/Users/felixdelorme/ESISAR/Stage/stage_felix_2024_MRR_Bolivie/MK_packaged_v2/Dat processing/precip_albedo_figures', 
    '/Users/felixdelorme/ESISAR/Stage/stage_felix_2024_MRR_Bolivie/MK_packaged_v2/Dat processing/precip_temp_figures', 
    '/Users/felixdelorme/ESISAR/Stage/stage_felix_2024_MRR_Bolivie/MK_packaged_v2/Dat processing/albedo_figures'
] 

excel_file = 'output.xlsx'

create_excel_file(directory, additional_directories, excel_file)
